import openpyxl
import time
import os

#changing directory to the location below
os.chdir('C:\\Users\\masly\\OneDrive\\Desktop\\School\\Semester 3\\CPT 180\\Labs')

#load the excel spreadsheet that program will be updating
workbook_file = openpyxl.load_workbook('Entrance Test Scores.xlsx')
sheet = workbook_file['Sheet1']

#Score update dictionary
score_update = {'AAS.AOT': 66, 'CT.COM': 44, 'CT.PPL': 44}

#Update score in sheet
for rowNum in range(2, sheet.max_row + 1):
     datael_acronym = sheet.cell(row = rowNum, column = 2).value
     if datael_acronym in score_update:
        sheet.cell(row = rowNum, column = 6).value = score_update[datael_acronym]

#save the update workboot
workbook_file.save('updateEntrance Test Scores.xlsx')

print('Scores updated successfully')

